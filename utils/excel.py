from openpyxl.styles import Alignment, Border, Side
from openpyxl.workbook import Workbook


def save_forecast_2_excel(weather_data):
    wb = Workbook()
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.merge_cells("B1:F1")
    ws["B1"] = "Прогноз погоды"
    ws["B1"].alignment = Alignment(horizontal='center')

    period_headers = ["Температура", "Влажность", "Давление", "Явление"]

    current_row = 3

    for day_data in weather_data:
        start_row = current_row

        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws[f"B{current_row}"] = "Дата"
        ws.merge_cells(f"D{current_row}:F{current_row}")
        ws[f"D{current_row}"] = day_data["date"]
        current_row += 1

        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws[f"B{current_row}"] = "Средняя температура"
        ws.merge_cells(f"D{current_row}:F{current_row}")
        ws[f"D{current_row}"] = day_data["day_average_temp"]
        current_row += 1

        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws[f"B{current_row}"] = "Магнитное поле"
        ws.merge_cells(f"D{current_row}:F{current_row}")
        ws[f"D{current_row}"] = day_data["magnetic_field_forecast"]
        current_row += 1

        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws[f"B{current_row}"] = "Предупреждение"
        ws.merge_cells(f"D{current_row}:F{current_row}")
        ws[f"D{current_row}"] = day_data["pressure_warning"]
        current_row += 1

        ws.merge_cells(f"B{current_row}:B{current_row + 11}")
        ws[f"B{current_row}"] = "Периоды"
        ws[f"B{current_row}"].alignment = Alignment(vertical='center', horizontal='center')

        periods = ["morning", "day", "evening", "night"]
        period_names = {"morning": "Утро", "day": "День", "evening": "Вечер", "night": "Ночь"}

        for period in periods:
            ws.merge_cells(f"C{current_row}:F{current_row}")
            ws[f"C{current_row}"] = period_names[period]
            current_row += 1

            for col, header in enumerate(period_headers, start=3):
                ws.cell(row=current_row, column=col, value=header)
            current_row += 1

            period_data = day_data["day_periods"][period]
            ws.cell(row=current_row, column=3, value=period_data["temperature_avg"])
            ws.cell(row=current_row, column=4, value=period_data["humidity"])
            ws.cell(row=current_row, column=5, value=period_data["pressure"])
            ws.cell(row=current_row, column=6, value=period_data["condition"])
            current_row += 1

        current_row += 1
        end_row = current_row - 1

        for row in range(start_row, end_row + 1):
            for col in range(2, 7):
                ws.cell(row=row, column=col).border = thin_border

        periods_cell = ws[f"B{start_row + 4}"]
        periods_cell.border = thin_border


    wb.save("weather_forecast.xlsx")

